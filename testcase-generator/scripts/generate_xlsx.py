"""
将测试用例 JSON 导出为 Excel。

支持两种模式：
1. python generate_xlsx.py <all_cases.json> <output.xlsx>
2. python generate_xlsx.py --merge <modules_dir> <output.xlsx>
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print(json.dumps({
        "status": "error",
        "message": "openpyxl 未安装，请先执行: pip install openpyxl",
    }, ensure_ascii=False))
    sys.exit(1)


COLUMNS = [
    ("功能模块", 18),
    ("用例编号", 14),
    ("功能点", 18),
    ("用例标题", 28),
    ("优先级", 10),
    ("前置条件", 28),
    ("测试步骤", 42),
    ("预期结果", 42),
    ("设计人", 12),
    ("执行结果", 12),
    ("执行人", 12),
    ("备注", 20),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", fgColor="F7FBFF")
NORMAL_FILL = PatternFill("solid", fgColor="FFFFFF")

PRIORITY_FILL = {
    "High": PatternFill("solid", fgColor="F4CCCC"),
    "Medium": PatternFill("solid", fgColor="FFF2CC"),
    "Low": PatternFill("solid", fgColor="D9EAD3"),
}


def style_cell(cell, fill, font=None, alignment=None):
    cell.fill = fill
    cell.font = font or BODY_FONT
    cell.alignment = alignment or BODY_ALIGN
    cell.border = BORDER


def normalize_case(case):
    normalized = {}
    for name, _width in COLUMNS:
        normalized[name] = str(case.get(name, "") or "")
    return normalized


def load_cases(path):
    data = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    if isinstance(data, list):
        return [normalize_case(item) for item in data]
    if isinstance(data, dict):
        cases = data.get("test_cases", data.get("cases", []))
        return [normalize_case(item) for item in cases]
    return []


def merge_cases(modules_dir):
    modules_path = Path(modules_dir)
    if not modules_path.exists() or not modules_path.is_dir():
        raise FileNotFoundError(f"目录不存在: {modules_dir}")

    all_cases = []
    seen = set()
    warnings = []
    for json_file in sorted(modules_path.glob("*.json")):
        if json_file.name.lower() == "all_cases.json":
            continue
        try:
            cases = load_cases(json_file)
        except json.JSONDecodeError as exc:
            warnings.append(f"JSON 解析失败 [{json_file.name}]: {exc}")
            continue

        for case in cases:
            case_id = case.get("用例编号", "")
            if case_id and case_id in seen:
                warnings.append(f"重复用例编号 [{case_id}]，已跳过后续重复项")
                continue
            if case_id:
                seen.add(case_id)
            all_cases.append(case)

    return all_cases, warnings


def write_excel(cases, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    for col_index, (column_name, _width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=column_name)
        style_cell(cell, HEADER_FILL, HEADER_FONT, HEADER_ALIGN)

    ws.row_dimensions[1].height = 28

    for row_index, raw_case in enumerate(cases, start=2):
        case = normalize_case(raw_case)
        base_fill = ALT_FILL if row_index % 2 == 0 else NORMAL_FILL

        for col_index, (column_name, _width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=case[column_name])
            style_cell(cell, base_fill)

        priority = case.get("优先级", "")
        if priority in PRIORITY_FILL:
            cell = ws.cell(row=row_index, column=5)
            style_cell(cell, PRIORITY_FILL[priority], BODY_FONT, BODY_ALIGN)

    for col_index, (column_name, default_width) in enumerate(COLUMNS, start=1):
        max_width = len(column_name)
        for row_index in range(2, len(cases) + 2):
            value = ws.cell(row=row_index, column=col_index).value
            if value:
                width = max(len(line) for line in str(value).split("\n"))
                max_width = max(max_width, width)
        adjusted = max(default_width, min(max_width * 1.2, default_width * 1.8))
        ws.column_dimensions[get_column_letter(col_index)].width = round(adjusted, 1)

    ws.freeze_panes = "A2"

    if len(cases) >= 1:
        last_col = get_column_letter(len(COLUMNS))
        last_row = len(cases) + 1
        table = Table(ref=f"A1:{last_col}{last_row}", displayName="TestCases")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return str(output.resolve())


def main():
    args = sys.argv[1:]
    if not args:
        print(json.dumps({
            "status": "error",
            "message": "用法: python generate_xlsx.py <all_cases.json> <output.xlsx> 或 python generate_xlsx.py --merge <modules_dir> <output.xlsx>",
        }, ensure_ascii=False))
        sys.exit(1)

    warnings = []
    if args[0] == "--merge":
        if len(args) != 3:
            print(json.dumps({
                "status": "error",
                "message": "用法: python generate_xlsx.py --merge <modules_dir> <output.xlsx>",
            }, ensure_ascii=False))
            sys.exit(1)
        modules_dir = args[1]
        output_path = args[2]
        try:
            cases, warnings = merge_cases(modules_dir)
        except FileNotFoundError as exc:
            print(json.dumps({"status": "error", "message": str(exc)}, ensure_ascii=False))
            sys.exit(1)
        all_cases_path = Path(output_path).parent / "all_cases.json"
        all_cases_path.write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        if len(args) != 2:
            print(json.dumps({
                "status": "error",
                "message": "用法: python generate_xlsx.py <all_cases.json> <output.xlsx>",
            }, ensure_ascii=False))
            sys.exit(1)
        data_path = Path(args[0])
        output_path = args[1]
        if not data_path.exists():
            print(json.dumps({
                "status": "error",
                "message": f"数据文件不存在: {data_path}",
            }, ensure_ascii=False))
            sys.exit(1)
        cases = load_cases(data_path)

    xlsx_path = write_excel(cases, output_path)
    result = {
        "status": "ok",
        "path": xlsx_path,
        "count": len(cases),
    }
    if args[0] == "--merge":
        result["all_cases_json"] = str((Path(output_path).parent / "all_cases.json").resolve())
    if warnings:
        result["warnings"] = warnings
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
